"""Comprehensive PI Planner parser for multi-sheet Excel workbooks."""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from pi_strategist.models import CapacityPlan, Sprint, Task

logger = logging.getLogger(__name__)


def normalize_discipline(discipline: str) -> str:
    """Normalize discipline names for consistent grouping.

    Examples:
        "PM Lead" -> "Project Management"
        "BE Group Lead" -> "Backend Engineering"
        "BA 1 Lead" -> "Business Analysis"
    """
    if not discipline:
        return "Other"

    disc_lower = discipline.lower().strip()

    # Project Management
    if disc_lower.startswith("pm") or "project manage" in disc_lower or "program" in disc_lower:
        return "Project Management"

    # Business Analysis
    if disc_lower.startswith("ba") or "business" in disc_lower or "analyst" in disc_lower:
        return "Business Analysis"

    # Backend Engineering
    if disc_lower.startswith("be") or "backend" in disc_lower or "back end" in disc_lower or "back-end" in disc_lower:
        return "Backend Engineering"

    # Frontend Engineering
    if disc_lower.startswith("fe") or "frontend" in disc_lower or "front end" in disc_lower or "front-end" in disc_lower or "ui" in disc_lower:
        return "Frontend Engineering"

    # QA / Testing
    if disc_lower.startswith("qa") or "quality" in disc_lower or "test" in disc_lower or "sdet" in disc_lower:
        return "Quality Assurance"

    # DevOps / Infrastructure
    if "devops" in disc_lower or "infra" in disc_lower or "platform" in disc_lower or "sre" in disc_lower:
        return "DevOps"

    # Design / UX
    if "design" in disc_lower or "ux" in disc_lower or "ui/ux" in disc_lower:
        return "Design"

    # Technical Writing / Documentation
    if "writer" in disc_lower or "documentation" in disc_lower or "technical writing" in disc_lower:
        return "Technical Writing"

    # Data / Analytics
    if "data" in disc_lower or "analytics" in disc_lower or "bi " in disc_lower:
        return "Data & Analytics"

    # Architecture
    if "architect" in disc_lower:
        return "Architecture"

    # Leadership / Management (catch-all for leads/managers)
    if "lead" in disc_lower or "manager" in disc_lower or "director" in disc_lower:
        # Try to extract the base discipline
        if "product" in disc_lower:
            return "Product"
        return "Leadership"

    # Product
    if "product" in disc_lower:
        return "Product"

    return discipline.title()  # Return original with title case


@dataclass
class Resource:
    """Represents a team member/resource."""
    name: str
    discipline: str = ""
    rate: float = 0.0
    total_hours: float = 0.0
    max_hours: float = 488.0  # PI maximum hours per person
    allocation_percentage: float = 0.0  # Target 100%
    sprint_hours: dict[str, float] = field(default_factory=dict)
    sprint_remaining: dict[str, float] = field(default_factory=dict)
    pto_hours: dict[str, float] = field(default_factory=dict)
    project_hours: dict[str, float] = field(default_factory=dict)

    @property
    def status(self) -> str:
        """Return allocation status: over, under, or optimal."""
        pct = (self.total_hours / self.max_hours * 100) if self.max_hours > 0 else 0
        if pct > 105:
            return "over"
        elif pct < 80:
            return "under"
        return "optimal"


@dataclass
class Project:
    """Represents a project in the PI."""
    name: str
    priority: int = 0
    description: str = ""
    total_hours: float = 0.0
    sprint_allocation: dict[str, bool] = field(default_factory=dict)  # Which sprints
    resource_hours: dict[str, float] = field(default_factory=dict)  # Resource -> hours


@dataclass
class Release:
    """Represents a release in the tracker."""
    name: str
    description: str = ""
    environment: str = ""
    staging_date: Optional[str] = None
    production_date: Optional[str] = None
    comments: str = ""


@dataclass
class PIAnalysis:
    """Complete PI analysis result."""
    sprints: dict[str, dict] = field(default_factory=dict)  # Sprint name -> {capacity, dates, etc}
    resources: dict[str, Resource] = field(default_factory=dict)
    projects: dict[str, Project] = field(default_factory=dict)
    releases: list[Release] = field(default_factory=list)

    # Validation results
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    # Cross-validation metrics
    total_capacity: float = 0.0
    total_allocated: float = 0.0
    total_remaining: float = 0.0
    overallocated_resources: list[tuple[str, str, float]] = field(default_factory=list)  # (resource, sprint, hours)

    # Grand totals from spreadsheet (read from Total row at bottom)
    grand_total_hours: float = 0.0  # Total Hours from spreadsheet's total row


class PIPlannerParser:
    """Parser for comprehensive PI planning Excel workbooks with multiple sheets."""

    def __init__(self, default_buffer: float = 0.20):
        self.default_buffer = default_buffer
        self._workbook = None

    def parse(self, file_path: str | Path) -> CapacityPlan:
        """Parse a PI planner Excel workbook.

        Args:
            file_path: Path to the Excel file

        Returns:
            Parsed CapacityPlan with cross-validated data
        """
        import openpyxl

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        self._workbook = openpyxl.load_workbook(path, data_only=True)

        # Analyze all sheets
        analysis = self._analyze_all_sheets()

        # Cross-validate data
        self._cross_validate(analysis)

        # Convert to CapacityPlan
        plan = self._to_capacity_plan(analysis, path.name)

        return plan

    def parse_with_analysis(self, file_path: str | Path) -> tuple[CapacityPlan, PIAnalysis]:
        """Parse and return both the CapacityPlan and detailed analysis."""
        import openpyxl

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        self._workbook = openpyxl.load_workbook(path, data_only=True)

        analysis = self._analyze_all_sheets()
        self._cross_validate(analysis)
        plan = self._to_capacity_plan(analysis, path.name)

        return plan, analysis

    def _analyze_all_sheets(self) -> PIAnalysis:
        """Analyze all sheets in the workbook."""
        analysis = PIAnalysis()

        for sheet_name in self._workbook.sheetnames:
            sheet = self._workbook[sheet_name]
            name_lower = sheet_name.lower()

            if "roadmap" in name_lower:
                self._parse_roadmap(sheet, analysis)
            elif "remaining" in name_lower:
                self._parse_remaining_hours(sheet, analysis)
            elif "project hours" in name_lower or (name_lower == "project hours"):
                self._parse_project_hours(sheet, analysis)
            # Handle truncated sheet names like "Worksheet for Resource Allocati"
            elif ("resource" in name_lower and "alloca" in name_lower) or "resource allocat" in name_lower:
                self._parse_resource_allocation(sheet, analysis)
            elif "percentage" in name_lower or "percent" in name_lower:
                self._parse_percentage(sheet, analysis)
            elif "per month" in name_lower or "monthly" in name_lower:
                self._parse_monthly_hours(sheet, analysis)
            elif "pto" in name_lower:
                self._parse_pto(sheet, analysis)
            elif "release" in name_lower or "tracker" in name_lower:
                self._parse_releases(sheet, analysis)

        return analysis

    def _parse_roadmap(self, sheet, analysis: PIAnalysis):
        """Parse the roadmap sheet for project-sprint mapping."""
        # Find sprint columns in header rows
        sprint_cols = {}
        priority_col = None
        description_col = None
        date_row = None

        for row_idx in range(1, min(10, sheet.max_row + 1)):
            for col_idx in range(1, min(30, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    cell_lower = cell_str.lower()

                    # Match "Sprint N" patterns
                    sprint_match = re.match(r'sprint\s*(\d+)', cell_lower)
                    if sprint_match:
                        sprint_num = sprint_match.group(1)
                        sprint_name = f"Sprint {sprint_num}"
                        sprint_cols[col_idx] = sprint_name
                        if sprint_name not in analysis.sprints:
                            analysis.sprints[sprint_name] = {"capacity": 0, "projects": [], "date_range": ""}
                    elif cell_lower == "priority":
                        priority_col = col_idx
                    elif cell_lower == "description":
                        description_col = col_idx
                    # Check for date ranges (e.g., "1/1-1/20" or "1/1 to 1/20")
                    elif re.match(r'\d{1,2}/\d{1,2}[-–]\d{1,2}/\d{1,2}', cell_str):
                        date_row = row_idx
                        # Find which sprint column this date belongs to
                        for sprint_col, sprint_name in sprint_cols.items():
                            if col_idx == sprint_col and sprint_name in analysis.sprints:
                                analysis.sprints[sprint_name]["date_range"] = cell_str

        if not sprint_cols:
            return

        # Parse project rows
        for row_idx in range(5, sheet.max_row + 1):
            # Get priority
            priority = 0
            if priority_col:
                prio_val = sheet.cell(row=row_idx, column=priority_col).value
                if prio_val:
                    try:
                        priority = int(prio_val)
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric priority value '%s': %s", prio_val, exc)

            # Get description
            description = ""
            if description_col:
                desc_val = sheet.cell(row=row_idx, column=description_col).value
                if desc_val:
                    description = str(desc_val).strip()

            if not description or priority == 0:
                continue

            # Check which sprints this project is in
            project_sprints = {}
            for col_idx, sprint_name in sprint_cols.items():
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                # Non-empty cell means project is in this sprint
                if cell_val is not None and str(cell_val).strip():
                    project_sprints[sprint_name] = True
                    if sprint_name in analysis.sprints:
                        if description not in analysis.sprints[sprint_name]["projects"]:
                            analysis.sprints[sprint_name]["projects"].append(description)

            if description not in analysis.projects:
                analysis.projects[description] = Project(
                    name=description,
                    priority=priority,
                    sprint_allocation=project_sprints,
                )

    def _parse_remaining_hours(self, sheet, analysis: PIAnalysis):
        """Parse remaining hours sheet for sprint capacity and resource availability."""
        # Find sprint columns and capacity row
        sprint_cols = {}
        header_row = None
        resource_col = None

        for row_idx in range(1, min(10, sheet.max_row + 1)):
            for col_idx in range(1, min(20, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    cell_lower = cell_str.lower()

                    if cell_lower.startswith("sprint"):
                        sprint_cols[col_idx] = cell_str
                    elif "total hours" in cell_lower or "available" in cell_lower:
                        # This row has capacity data
                        for sc_col, sc_name in sprint_cols.items():
                            cap_val = sheet.cell(row=row_idx, column=sc_col).value
                            if cap_val:
                                try:
                                    capacity = float(cap_val)
                                    if sc_name not in analysis.sprints:
                                        analysis.sprints[sc_name] = {"capacity": 0, "projects": []}
                                    analysis.sprints[sc_name]["capacity"] = capacity
                                    analysis.total_capacity += capacity
                                except (ValueError, TypeError) as exc:
                                    logger.warning("Non-numeric sprint capacity for %s: %s", sc_name, exc)
                    elif cell_lower in ["team member", "resource", "name"]:
                        resource_col = col_idx
                        header_row = row_idx

        if not (resource_col and sprint_cols and header_row):
            return

        # Parse resource remaining hours
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            resource_name = sheet.cell(row=row_idx, column=resource_col).value
            if not resource_name:
                continue
            resource_name = str(resource_name).strip()

            # Get discipline if available and normalize it
            discipline = ""
            if resource_col > 1:
                disc_val = sheet.cell(row=row_idx, column=resource_col - 1).value
                if disc_val:
                    discipline = normalize_discipline(str(disc_val).strip())

            if resource_name not in analysis.resources:
                analysis.resources[resource_name] = Resource(name=resource_name, discipline=discipline)

            for col_idx, sprint_name in sprint_cols.items():
                hours_val = sheet.cell(row=row_idx, column=col_idx).value
                if hours_val is not None:
                    try:
                        hours = float(hours_val)
                        analysis.resources[resource_name].sprint_remaining[sprint_name] = hours
                        analysis.total_remaining += hours

                        # Track overallocation
                        if hours < 0:
                            analysis.overallocated_resources.append(
                                (resource_name, sprint_name, abs(hours))
                            )
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric remaining hours for %s in %s: %s", resource_name, sprint_name, exc)

    def _parse_project_hours(self, sheet, analysis: PIAnalysis):
        """Parse project hours sheet for project-resource-hours mapping."""
        # Find header row with project names and Hours columns
        project_cols = {}  # col_idx -> project_name
        header_row = None
        resource_col = None
        rate_col = None
        discipline_col = None

        # First pass: find project names row (usually has project names spanning multiple columns)
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            current_project = None
            for col_idx in range(1, min(60, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    # Skip common non-project headers
                    if cell_str.lower() in ["none", "max", "projects", "total", "delta"]:
                        continue
                    if len(cell_str) > 5 and not cell_str.lower().startswith(("hour", "price", "rate")):
                        current_project = cell_str[:50]

                # Check next row for "Hours" header
                if current_project:
                    next_row_val = sheet.cell(row=row_idx + 1, column=col_idx).value
                    if next_row_val and str(next_row_val).lower().startswith("hour"):
                        project_cols[col_idx] = current_project
                        if current_project not in analysis.projects:
                            analysis.projects[current_project] = Project(name=current_project)

        # Find resource column, rate column, and header row
        for row_idx in range(1, min(15, sheet.max_row + 1)):
            for col_idx in range(1, min(10, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).lower().strip()
                    if cell_str in ["resource", "name", "team member"]:
                        resource_col = col_idx
                        header_row = row_idx
                    elif cell_str == "rate":
                        rate_col = col_idx
                    elif cell_str in ["column1", "discipline", "role"]:
                        discipline_col = col_idx
                    elif cell_str in ["mqd as", "title", "job title"]:
                        # This is the job title/discipline column
                        if not discipline_col:
                            discipline_col = col_idx
            if header_row and rate_col:
                break

        if not (resource_col and project_cols and header_row):
            return

        # Parse resource-project hours
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            resource_name = sheet.cell(row=row_idx, column=resource_col).value
            if not resource_name:
                continue
            resource_name = str(resource_name).strip()

            # Skip header-like rows
            if resource_name.lower() in ["resource", "name", "total"]:
                continue

            if resource_name not in analysis.resources:
                analysis.resources[resource_name] = Resource(name=resource_name)

            # Get rate
            if rate_col:
                rate_val = sheet.cell(row=row_idx, column=rate_col).value
                if rate_val is not None:
                    try:
                        analysis.resources[resource_name].rate = float(rate_val)
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric rate for %s: %s", resource_name, exc)

            # Get discipline/role and normalize it
            if discipline_col:
                disc_val = sheet.cell(row=row_idx, column=discipline_col).value
                if disc_val:
                    analysis.resources[resource_name].discipline = normalize_discipline(str(disc_val).strip())

            for col_idx, project_name in project_cols.items():
                hours_val = sheet.cell(row=row_idx, column=col_idx).value
                if hours_val is not None:
                    try:
                        hours = float(hours_val)
                        if hours > 0:
                            analysis.resources[resource_name].project_hours[project_name] = hours
                            # Don't accumulate total_hours here - it will be set from
                            # the Total Hours column in _parse_resource_allocation

                            # Update project totals
                            if project_name in analysis.projects:
                                analysis.projects[project_name].resource_hours[resource_name] = hours
                                analysis.projects[project_name].total_hours += hours

                            # Don't accumulate total_allocated - it's recalculated in _cross_validate
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric project hours for %s / %s: %s", resource_name, project_name, exc)

    def _parse_resource_allocation(self, sheet, analysis: PIAnalysis):
        """Parse resource allocation worksheet for hours per person per project per sprint."""
        # Find header row with Discipline, Team Member, etc.
        header_row = None
        discipline_col = None
        resource_col = None
        role_col = None
        total_alloc_col = None
        total_hours_col = None
        sprint_cols = {}  # col_idx -> (project_name, sprint_name)

        # First pass: find the main header row
        for row_idx in range(1, min(15, sheet.max_row + 1)):
            for col_idx in range(1, min(15, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip().lower()
                    if cell_str == "discipline":
                        discipline_col = col_idx
                        header_row = row_idx
                    elif cell_str in ["team member", "resource", "name"]:
                        resource_col = col_idx
                    elif cell_str in ["billable role", "role", "title"]:
                        role_col = col_idx
                    elif "total allocation" in cell_str:
                        total_alloc_col = col_idx
                    elif "total hours" in cell_str:
                        total_hours_col = col_idx

            if header_row:
                break

        if not (header_row and resource_col):
            return

        # Second pass: find project names and sprint columns
        # Projects are in the row above header, sprints are in header row
        project_row = header_row - 3 if header_row > 3 else 1
        sprint_row = header_row - 1 if header_row > 1 else header_row

        current_project = None
        for col_idx in range(8, min(130, sheet.max_column + 1)):  # Start after main columns
            # Get project name from project row
            project_cell = sheet.cell(row=project_row, column=col_idx).value
            if project_cell:
                project_str = str(project_cell).strip()
                if len(project_str) > 5 and project_str.lower() not in ["project", "total"]:
                    current_project = project_str[:50]
                    # Create project if not exists
                    if current_project not in analysis.projects:
                        analysis.projects[current_project] = Project(name=current_project)

            # Get sprint name from sprint row
            sprint_cell = sheet.cell(row=sprint_row, column=col_idx).value
            if sprint_cell and current_project:
                sprint_str = str(sprint_cell).strip()
                # Match Sprint N or month names
                if sprint_str.lower().startswith("sprint") or sprint_str.lower() in [
                    "jan", "feb", "mar", "apr", "may", "jun",
                    "jul", "aug", "sep", "oct", "nov", "dec",
                    "jan/feb", "feb/mar", "mar/apr"
                ]:
                    sprint_cols[col_idx] = (current_project, sprint_str)

        # Parse resource rows
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            # Get resource name
            resource_name = sheet.cell(row=row_idx, column=resource_col).value
            if not resource_name:
                continue
            resource_name = str(resource_name).strip()

            # Skip header-like rows
            if resource_name.lower() in ["team member", "resource", "discipline", "sprint hours"]:
                continue

            # Check for Total row - capture grand total and skip
            if resource_name.lower() in ["total", "grand total", "totals"]:
                if total_hours_col:
                    total_val = sheet.cell(row=row_idx, column=total_hours_col).value
                    if total_val is not None:
                        try:
                            analysis.grand_total_hours = float(total_val)
                        except (ValueError, TypeError) as exc:
                            logger.warning("Non-numeric grand total hours: %s", exc)
                continue  # Skip adding Total as a resource

            # Get or create resource
            if resource_name not in analysis.resources:
                analysis.resources[resource_name] = Resource(name=resource_name)

            resource = analysis.resources[resource_name]

            # Get discipline and normalize it
            if discipline_col:
                disc_val = sheet.cell(row=row_idx, column=discipline_col).value
                if disc_val:
                    resource.discipline = normalize_discipline(str(disc_val).strip())

            # Get role (fallback if no discipline)
            if role_col:
                role_val = sheet.cell(row=row_idx, column=role_col).value
                if role_val and not resource.discipline:
                    resource.discipline = normalize_discipline(str(role_val).strip())

            # Get total allocation percentage
            if total_alloc_col:
                alloc_val = sheet.cell(row=row_idx, column=total_alloc_col).value
                if alloc_val is not None:
                    try:
                        pct = float(alloc_val)
                        if pct <= 2:  # Decimal format
                            pct = pct * 100
                        resource.allocation_percentage = pct
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric allocation % for %s: %s", resource.name, exc)

            # Get total hours
            if total_hours_col:
                hours_val = sheet.cell(row=row_idx, column=total_hours_col).value
                if hours_val is not None:
                    try:
                        resource.total_hours = float(hours_val)
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric total hours for %s: %s", resource.name, exc)

            # Get hours per project per sprint
            for col_idx, (project_name, sprint_name) in sprint_cols.items():
                hours_val = sheet.cell(row=row_idx, column=col_idx).value
                if hours_val is not None:
                    try:
                        hours = float(hours_val)
                        if hours > 0:
                            # Track sprint hours (this is unique to this sheet)
                            if sprint_name not in resource.sprint_hours:
                                resource.sprint_hours[sprint_name] = 0
                            resource.sprint_hours[sprint_name] += hours

                            # Don't accumulate project_hours here - it's already set from _parse_project_hours
                            # Don't accumulate project totals - already done in _parse_project_hours

                            # Mark project as active in this sprint
                            normalized_sprint = sprint_name
                            if sprint_name.lower().startswith("sprint"):
                                normalized_sprint = sprint_name
                            elif "/" in sprint_name:  # Month format like "Jan/Feb"
                                # Map to sprint number based on position
                                pass

                            if project_name in analysis.projects:
                                analysis.projects[project_name].sprint_allocation[normalized_sprint] = True

                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric sprint hours for %s / %s: %s", resource.name, project_name, exc)

    def _parse_percentage(self, sheet, analysis: PIAnalysis):
        """Parse percentage allocation sheet for resource allocation percentages."""
        # Find header row with resource names and project columns
        header_row = None
        resource_col = None
        total_alloc_col = None
        project_cols = {}  # col_idx -> project_name

        for row_idx in range(1, min(10, sheet.max_row + 1)):
            for col_idx in range(1, min(80, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    cell_lower = cell_str.lower()

                    if cell_lower in ["team member", "resource", "name"]:
                        resource_col = col_idx
                        header_row = row_idx
                    elif "total allocation" in cell_lower or cell_lower == "total allocation":
                        total_alloc_col = col_idx
                    elif "only dd&i" in cell_lower or cell_lower == "only dd&i":
                        # This is the DD&I-only allocation column
                        pass
                    # Check for project name patterns (long strings that aren't standard headers)
                    elif len(cell_str) > 10 and cell_lower not in ["discipline", "total allocation"]:
                        # Likely a project name
                        project_cols[col_idx] = cell_str[:50]

        if not (resource_col and header_row):
            return

        # Parse percentage values per resource
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            resource_name = sheet.cell(row=row_idx, column=resource_col).value
            if not resource_name:
                continue
            resource_name = str(resource_name).strip()

            # Skip header-like rows
            if resource_name.lower() in ["resource", "name", "team member", "total"]:
                continue

            # Get or create resource
            if resource_name not in analysis.resources:
                analysis.resources[resource_name] = Resource(name=resource_name)

            # Get total allocation percentage
            if total_alloc_col:
                alloc_val = sheet.cell(row=row_idx, column=total_alloc_col).value
                if alloc_val is not None:
                    try:
                        # Convert decimal to percentage if needed (0.85 -> 85%)
                        pct = float(alloc_val)
                        if pct <= 2:  # Likely a decimal
                            pct = pct * 100
                        analysis.resources[resource_name].allocation_percentage = pct
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric total allocation for %s: %s", resource_name, exc)

            # Also sum up project allocations
            total_project_pct = 0.0
            for col_idx, project_name in project_cols.items():
                pct_val = sheet.cell(row=row_idx, column=col_idx).value
                if pct_val is not None:
                    try:
                        pct = float(pct_val)
                        if pct <= 2:  # Likely a decimal
                            pct = pct * 100
                        total_project_pct += pct
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric project allocation for %s / %s: %s", resource_name, project_name, exc)

            # If no total allocation col, use summed project percentages
            if not total_alloc_col and total_project_pct > 0:
                analysis.resources[resource_name].allocation_percentage = total_project_pct

    def _parse_monthly_hours(self, sheet, analysis: PIAnalysis):
        """Parse monthly hours breakdown."""
        # Find header row
        header_row = None
        resource_col = None
        month_cols = {}

        for row_idx in range(1, min(5, sheet.max_row + 1)):
            for col_idx in range(1, min(20, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip().lower()
                    if cell_str == "resource":
                        resource_col = col_idx
                        header_row = row_idx
                    elif cell_str in ["january", "february", "march", "april", "may", "june",
                                      "july", "august", "september", "october", "november", "december",
                                      "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]:
                        month_cols[col_idx] = cell_str

        if not (resource_col and header_row):
            return

        # Parse resource monthly hours
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            resource_name = sheet.cell(row=row_idx, column=resource_col).value
            if not resource_name:
                continue
            resource_name = str(resource_name).strip()

            if resource_name not in analysis.resources:
                analysis.resources[resource_name] = Resource(name=resource_name)

    def _parse_pto(self, sheet, analysis: PIAnalysis):
        """Parse PTO analysis sheet."""
        # Find header row with resource names
        header_row = None
        resource_col = None
        total_col = None

        for row_idx in range(1, min(10, sheet.max_row + 1)):
            for col_idx in range(1, min(10, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip().lower()
                    if cell_str in ["team member", "resource", "name"]:
                        resource_col = col_idx
                        header_row = row_idx
                    elif cell_str == "total":
                        total_col = col_idx

        if not (resource_col and header_row):
            return

        # Parse PTO hours
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            resource_name = sheet.cell(row=row_idx, column=resource_col).value
            if not resource_name:
                continue
            resource_name = str(resource_name).strip()

            if resource_name not in analysis.resources:
                analysis.resources[resource_name] = Resource(name=resource_name)

            # Get total PTO
            if total_col:
                pto_val = sheet.cell(row=row_idx, column=total_col).value
                if pto_val:
                    try:
                        pto_hours = float(pto_val)
                        analysis.resources[resource_name].pto_hours["total"] = pto_hours
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric PTO hours for %s: %s", resource_name, exc)

    def _parse_releases(self, sheet, analysis: PIAnalysis):
        """Parse release tracker sheet."""
        # Find header row
        header_row = None
        cols = {}

        for row_idx in range(1, min(5, sheet.max_row + 1)):
            for col_idx in range(1, min(10, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip().lower()
                    if "release" in cell_str:
                        cols["release"] = col_idx
                        header_row = row_idx
                    elif "description" in cell_str:
                        cols["description"] = col_idx
                    elif "env" in cell_str or "test" in cell_str:
                        cols["environment"] = col_idx
                    elif "stg" in cell_str or "staging" in cell_str:
                        cols["staging"] = col_idx
                    elif "prd" in cell_str or "prod" in cell_str:
                        cols["production"] = col_idx
                    elif "comment" in cell_str:
                        cols["comments"] = col_idx

        if not header_row:
            return

        # Parse releases
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            release_name = ""
            if "release" in cols:
                rel_val = sheet.cell(row=row_idx, column=cols["release"]).value
                if rel_val:
                    release_name = str(rel_val).strip()

            if not release_name:
                continue

            release = Release(name=release_name)

            if "description" in cols:
                desc_val = sheet.cell(row=row_idx, column=cols["description"]).value
                if desc_val:
                    release.description = str(desc_val).strip()

            if "environment" in cols:
                env_val = sheet.cell(row=row_idx, column=cols["environment"]).value
                if env_val:
                    release.environment = str(env_val).strip()

            if "staging" in cols:
                stg_val = sheet.cell(row=row_idx, column=cols["staging"]).value
                if stg_val:
                    release.staging_date = str(stg_val)

            if "production" in cols:
                prd_val = sheet.cell(row=row_idx, column=cols["production"]).value
                if prd_val:
                    release.production_date = str(prd_val)

            if "comments" in cols:
                cmt_val = sheet.cell(row=row_idx, column=cols["comments"]).value
                if cmt_val:
                    release.comments = str(cmt_val).strip()

            analysis.releases.append(release)

    def _cross_validate(self, analysis: PIAnalysis):
        """Cross-validate data between sheets."""
        PI_MAX_HOURS = 488.0  # Maximum hours per person for the PI

        # Ensure each resource has total_hours set
        # If not set from spreadsheet, calculate from project_hours
        for resource in analysis.resources.values():
            if resource.total_hours == 0 and resource.project_hours:
                resource.total_hours = sum(resource.project_hours.values())

        # Remove blank/unassigned resources and invalid rows
        invalid_names = {"total", "grand total", "totals", "subtotal", "sum", "", "resource", "name", "team member"}
        resources_to_remove = [
            name for name, resource in analysis.resources.items()
            if (
                # No hours allocated
                (resource.total_hours == 0 and not resource.project_hours and not resource.sprint_hours)
                # Or name looks like a header/total row
                or name.lower().strip() in invalid_names
                # Or name is too short (likely invalid)
                or len(name.strip()) < 2
            )
        ]
        for name in resources_to_remove:
            del analysis.resources[name]

        # Correct the total_capacity and total_allocated values
        # total_capacity should be 488h per resource (standard PI capacity)
        num_resources = len(analysis.resources)
        analysis.total_capacity = num_resources * PI_MAX_HOURS

        # total_allocated should use grand_total_hours from spreadsheet if available,
        # otherwise sum the individual resource hours
        calculated_total = sum(r.total_hours for r in analysis.resources.values())

        if analysis.grand_total_hours > 0:
            analysis.total_allocated = analysis.grand_total_hours
        else:
            analysis.total_allocated = calculated_total

        # Check for overallocation
        if analysis.overallocated_resources:
            for resource, sprint, hours in analysis.overallocated_resources:
                analysis.warnings.append(
                    f"Resource '{resource}' is over-allocated by {hours:.1f}h in {sprint}"
                )

        # Compare total allocated vs capacity
        if analysis.total_capacity > 0:
            utilization = (analysis.total_allocated / analysis.total_capacity) * 100
            if utilization > 100:
                analysis.warnings.append(
                    f"Total allocation ({analysis.total_allocated:.1f}h) exceeds capacity "
                    f"({analysis.total_capacity:.1f}h) by {utilization - 100:.1f}%"
                )
            elif utilization < 50:
                analysis.warnings.append(
                    f"Low utilization: only {utilization:.1f}% of capacity is allocated"
                )

        # Check individual resource allocation vs PI max (488 hours)
        for resource_name, resource in analysis.resources.items():
            if resource.total_hours > PI_MAX_HOURS:
                over_by = resource.total_hours - PI_MAX_HOURS
                pct = (resource.total_hours / PI_MAX_HOURS) * 100
                analysis.warnings.append(
                    f"Resource '{resource_name}' is allocated {resource.total_hours:.0f}h "
                    f"({pct:.0f}%), exceeding PI max of {PI_MAX_HOURS:.0f}h by {over_by:.0f}h"
                )
            elif resource.total_hours > 0 and resource.total_hours < PI_MAX_HOURS * 0.80:
                pct = (resource.total_hours / PI_MAX_HOURS) * 100
                analysis.warnings.append(
                    f"Resource '{resource_name}' is under-allocated at {resource.total_hours:.0f}h "
                    f"({pct:.0f}% of {PI_MAX_HOURS:.0f}h target)"
                )

            # Calculate allocation percentage if not already set
            if resource.allocation_percentage == 0 and resource.total_hours > 0:
                resource.allocation_percentage = (resource.total_hours / PI_MAX_HOURS) * 100

        # Check for resources in projects but not in remaining hours
        for resource_name, resource in analysis.resources.items():
            if resource.project_hours and not resource.sprint_remaining:
                analysis.warnings.append(
                    f"Resource '{resource_name}' has project allocations but no sprint remaining hours"
                )

        # Check for projects in roadmap without hours
        for project_name, project in analysis.projects.items():
            if project.sprint_allocation and project.total_hours == 0:
                analysis.warnings.append(
                    f"Project '{project_name[:30]}...' is on roadmap but has no hours allocated"
                )

    def _to_capacity_plan(self, analysis: PIAnalysis, filename: str) -> CapacityPlan:
        """Convert PIAnalysis to CapacityPlan."""
        plan = CapacityPlan(filename=filename)

        # Calculate total hours per sprint from resource remaining data
        # remaining = available - allocated, so allocated = available - remaining
        sprint_allocated = {}
        for sprint_name in analysis.sprints:
            sprint_allocated[sprint_name] = 0.0

        for resource_name, resource in analysis.resources.items():
            for sprint_name, remaining in resource.sprint_remaining.items():
                if sprint_name in sprint_allocated:
                    # Negative remaining = over-allocated
                    # Positive remaining = available capacity left
                    # The amount allocated is (some base) - remaining
                    # If remaining is negative, they're over capacity
                    if remaining < 0:
                        sprint_allocated[sprint_name] += abs(remaining)

        # Create sprints
        for sprint_name, sprint_data in sorted(analysis.sprints.items()):
            capacity = sprint_data.get("capacity", 0)
            projects_in_sprint = sprint_data.get("projects", [])

            # Create tasks from projects in this sprint
            tasks = []
            task_idx = 1

            for project_name, project in analysis.projects.items():
                if project.sprint_allocation.get(sprint_name) or project_name in projects_in_sprint:
                    # Estimate hours for this sprint (divide by number of sprints)
                    num_sprints = len([s for s, v in project.sprint_allocation.items() if v])
                    if num_sprints == 0:
                        num_sprints = 1
                    sprint_hours = project.total_hours / num_sprints

                    if sprint_hours > 0:
                        tasks.append(Task(
                            id=f"{sprint_name.replace(' ', '')}-PROJ-{task_idx:03d}",
                            name=project_name[:50],
                            hours=sprint_hours,
                            sprint=sprint_name,
                            tags=["project", f"priority-{project.priority}"] if project.priority else ["project"],
                        ))
                        task_idx += 1

            # If no project-based tasks, calculate load from resource allocations
            if not tasks:
                # Group resources by discipline for cleaner task breakdown
                discipline_hours = {}
                for resource_name, resource in analysis.resources.items():
                    remaining = resource.sprint_remaining.get(sprint_name, None)
                    if remaining is not None:
                        discipline = resource.discipline or "General"
                        # Calculate hours used: if remaining < 0, over-allocated by that amount
                        # We estimate each person has ~capacity/num_resources base hours
                        per_person_base = capacity / max(len(analysis.resources), 1) * 4  # rough estimate
                        hours_used = per_person_base - remaining if remaining < per_person_base else 0

                        if hours_used > 0:
                            if discipline not in discipline_hours:
                                discipline_hours[discipline] = 0
                            discipline_hours[discipline] += hours_used

                for discipline, hours in discipline_hours.items():
                    if hours > 0:
                        tasks.append(Task(
                            id=f"{sprint_name.replace(' ', '')}-{discipline[:3].upper()}-{task_idx:03d}",
                            name=f"{discipline} Team Allocation",
                            hours=hours,
                            sprint=sprint_name,
                            tags=["resource-allocation", discipline.lower()],
                        ))
                        task_idx += 1

            # If still no tasks, add overallocated resources as warning tasks
            if not tasks:
                for resource_name, resource in analysis.resources.items():
                    remaining = resource.sprint_remaining.get(sprint_name, 0)
                    if remaining < 0:  # Over-allocated
                        tasks.append(Task(
                            id=f"{sprint_name.replace(' ', '')}-OVER-{task_idx:03d}",
                            name=f"{resource_name} (over-allocated)",
                            hours=abs(remaining),
                            sprint=sprint_name,
                            tags=["over-allocated", "warning"],
                        ))
                        task_idx += 1

            sprint = Sprint(
                name=sprint_name,
                total_hours=capacity if capacity > 0 else sum(t.hours for t in tasks) * 1.25,
                buffer_percentage=self.default_buffer,
                tasks=tasks,
            )
            plan.sprints.append(sprint)

        # Sort sprints by name
        plan.sprints.sort(key=lambda s: s.name)

        return plan
