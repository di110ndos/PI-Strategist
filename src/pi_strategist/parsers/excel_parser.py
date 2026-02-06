"""Parser for Excel capacity planners."""

import logging
import re
from pathlib import Path
from typing import Optional

from pi_strategist.models import CapacityPlan, Sprint, Task
from pi_strategist.parsers.pi_planner_parser import PIPlannerParser

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parser for Excel capacity planners."""

    # Common column name patterns
    TASK_ID_COLUMNS = ["task id", "task_id", "taskid", "id", "ticket", "jira"]
    TASK_NAME_COLUMNS = ["task", "task name", "name", "description", "title", "summary"]
    HOURS_COLUMNS = ["hours", "estimate", "estimated hours", "effort", "story points", "sp", "pts"]
    SPRINT_COLUMNS = ["sprint", "iteration", "pi", "release"]
    STORY_COLUMNS = ["story", "story id", "user story", "parent", "epic"]
    DEPENDENCY_COLUMNS = ["dependencies", "depends on", "blocked by", "blockers"]

    def __init__(self, default_buffer: float = 0.20):
        """Initialize the Excel parser.

        Args:
            default_buffer: Default buffer percentage (0.20 = 20%)
        """
        self.default_buffer = default_buffer
        self._openpyxl_available = self._check_openpyxl()

    def _check_openpyxl(self) -> bool:
        """Check if openpyxl is available."""
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    def parse(self, file_path: str | Path) -> CapacityPlan:
        """Parse a capacity planner from Excel file.

        Args:
            file_path: Path to the Excel file (.xlsx or .xls)

        Returns:
            Parsed CapacityPlan
        """
        if not self._openpyxl_available:
            raise ImportError("openpyxl is required to parse Excel files")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = path.suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            raise ValueError(f"Unsupported file format: {suffix}")

        return self._parse_excel(path)

    def _parse_excel(self, path: Path) -> CapacityPlan:
        """Parse an Excel workbook."""
        import openpyxl

        workbook = openpyxl.load_workbook(path, data_only=True)

        # First, try the comprehensive PI planner parser for multi-sheet workbooks
        if len(workbook.sheetnames) > 3:
            try:
                pi_parser = PIPlannerParser(default_buffer=self.default_buffer)
                plan = pi_parser.parse(path)
                if plan.sprints:
                    return plan
            except Exception as exc:
                logger.warning("PI planner parse failed, falling back to standard parsing: %s", exc)

        plan = CapacityPlan(filename=path.name)

        # Check if sheets are named as sprints
        sprint_sheets = self._identify_sprint_sheets(workbook.sheetnames)

        if sprint_sheets:
            # Each sheet is a sprint
            for sheet_name, sprint_name in sprint_sheets.items():
                sheet = workbook[sheet_name]
                sprint = self._parse_sprint_sheet(sheet, sprint_name)
                if sprint.tasks:
                    plan.sprints.append(sprint)
        else:
            # Try single sheet with sprint column first
            sheet = workbook.active
            if sheet:
                plan.sprints = self._parse_combined_sheet(sheet)

        # If no sprints found, try resource allocation format
        if not plan.sprints:
            plan.sprints = self._parse_resource_allocation(workbook)

        return plan

    def _parse_resource_allocation(self, workbook) -> list[Sprint]:
        """Parse resource allocation matrix format (projects as columns, resources as rows).

        This handles Excel files where:
        - Projects are listed as column headers
        - Resources/people are listed as rows
        - Hours are at the intersection
        """
        sprints = []

        # First, try to get sprint capacity from "Remaining Hours" or similar sheet
        sprint_capacity = self._extract_sprint_capacity(workbook)

        # Look for sheets that might contain project hours
        target_sheets = []
        for name in workbook.sheetnames:
            name_lower = name.lower()
            if any(kw in name_lower for kw in ["project hours", "hours", "allocation"]):
                target_sheets.append(name)

        # If no matching sheets, try first sheet
        if not target_sheets and workbook.sheetnames:
            target_sheets = [workbook.sheetnames[0]]

        for sheet_name in target_sheets:
            sheet = workbook[sheet_name]
            result = self._parse_resource_sheet(sheet, sheet_name, sprint_capacity)
            if result:
                sprints.extend(result)
                break  # Use first successful parse

        # If we have sprint capacity but no project tasks, create sprints from capacity data
        if not sprints and sprint_capacity:
            sprints = self._create_sprints_from_capacity(sprint_capacity, workbook)

        return sprints

    def _extract_sprint_capacity(self, workbook) -> dict[str, float]:
        """Extract sprint capacity from Remaining Hours or similar sheet."""
        sprint_capacity = {}

        # Look for remaining hours sheet
        for sheet_name in workbook.sheetnames:
            name_lower = sheet_name.lower()
            if any(kw in name_lower for kw in ["remaining", "capacity", "available"]):
                sheet = workbook[sheet_name]
                capacity = self._parse_capacity_sheet(sheet)
                if capacity:
                    sprint_capacity.update(capacity)
                    break

        return sprint_capacity

    def _parse_capacity_sheet(self, sheet) -> dict[str, float]:
        """Parse a capacity/remaining hours sheet to get sprint hours."""
        capacity = {}

        # Look for rows with "Total Hours" or similar and sprint columns
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            # Check if this row has sprint names
            sprint_cols = {}
            for col_idx in range(1, min(20, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    if cell_str.lower().startswith("sprint"):
                        sprint_cols[col_idx] = cell_str

            if sprint_cols:
                # Found sprint headers, look for hours in next row or same row pattern
                for check_row in [row_idx + 1, row_idx]:
                    # Check if this row has "hours" or "total" indicator
                    row_has_hours = False
                    for col_idx in range(1, 5):
                        cell_val = sheet.cell(row=check_row, column=col_idx).value
                        if cell_val and any(kw in str(cell_val).lower() for kw in ["hour", "total", "available", "capacity"]):
                            row_has_hours = True
                            break

                    if row_has_hours or check_row == row_idx + 1:
                        # Extract hours for each sprint
                        for col_idx, sprint_name in sprint_cols.items():
                            hours_cell = sheet.cell(row=check_row, column=col_idx).value
                            if hours_cell is not None:
                                try:
                                    hours = float(hours_cell)
                                    if hours > 0:
                                        capacity[sprint_name] = hours
                                except (ValueError, TypeError) as exc:
                                    logger.warning("Non-numeric capacity value in %s: %s", sprint_name, exc)
                        if capacity:
                            return capacity

        return capacity

    def _create_sprints_from_capacity(self, sprint_capacity: dict[str, float], workbook) -> list[Sprint]:
        """Create Sprint objects from capacity data, pulling in resource allocation."""
        sprints = []

        # Get resource allocation data if available
        resource_hours = self._extract_resource_allocation(workbook, list(sprint_capacity.keys()))

        for sprint_name, total_hours in sorted(sprint_capacity.items()):
            # Create tasks from resource allocations for this sprint
            tasks = []
            if sprint_name in resource_hours:
                for i, (resource, hours) in enumerate(resource_hours[sprint_name].items(), 1):
                    if hours != 0:  # Include negative (over-allocated) too
                        tasks.append(Task(
                            id=f"{sprint_name.replace(' ', '')}-{i:03d}",
                            name=f"{resource} allocation",
                            hours=abs(hours),  # Use absolute for task hours
                            sprint=sprint_name,
                            tags=["resource-allocation", "over-allocated" if hours < 0 else "available"],
                        ))

            sprint = Sprint(
                name=sprint_name,
                total_hours=total_hours,
                buffer_percentage=self.default_buffer,
                tasks=tasks,
            )
            sprints.append(sprint)

        return sprints

    def _extract_resource_allocation(self, workbook, sprint_names: list[str]) -> dict[str, dict[str, float]]:
        """Extract resource hours per sprint from allocation sheets."""
        result = {s: {} for s in sprint_names}

        # Look for remaining hours or resource allocation sheet
        for sheet_name in workbook.sheetnames:
            name_lower = sheet_name.lower()
            if any(kw in name_lower for kw in ["remaining", "resource", "allocation"]):
                sheet = workbook[sheet_name]

                # Find header row with sprint columns
                header_row = None
                sprint_cols = {}
                resource_col = None

                for row_idx in range(1, min(15, sheet.max_row + 1)):
                    for col_idx in range(1, min(20, sheet.max_column + 1)):
                        cell_val = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_val:
                            cell_str = str(cell_val).strip()
                            if cell_str.lower() in ["team member", "resource", "name"]:
                                resource_col = col_idx
                                header_row = row_idx
                            for sprint in sprint_names:
                                if cell_str.lower() == sprint.lower():
                                    sprint_cols[sprint] = col_idx

                    if resource_col and sprint_cols:
                        break

                if not (resource_col and sprint_cols):
                    continue

                # Parse resource rows
                for row_idx in range(header_row + 1, sheet.max_row + 1):
                    resource_cell = sheet.cell(row=row_idx, column=resource_col).value
                    if not resource_cell:
                        continue
                    resource_name = str(resource_cell).strip()

                    for sprint_name, col_idx in sprint_cols.items():
                        hours_cell = sheet.cell(row=row_idx, column=col_idx).value
                        if hours_cell is not None:
                            try:
                                hours = float(hours_cell)
                                result[sprint_name][resource_name] = hours
                            except (ValueError, TypeError) as exc:
                                logger.warning("Non-numeric hours for %s in %s: %s", resource_name, sprint_name, exc)

                if any(result[s] for s in sprint_names):
                    break

        return result

    def _parse_resource_sheet(self, sheet, sheet_name: str, sprint_capacity: dict[str, float] = None) -> list[Sprint]:
        """Parse a resource allocation sheet."""
        # Find the header row with project names (look for row with "Hours" columns)
        project_cols = {}  # project_name -> column_index
        header_row = None
        resource_col = None

        for row_idx in range(1, min(15, sheet.max_row + 1)):
            row_has_hours = False
            row_has_resource = False

            for col_idx in range(1, min(50, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).lower().strip()
                    if cell_str in ("hours", "hrs", "hour"):
                        row_has_hours = True
                    if cell_str in ("resource", "name", "team member", "person"):
                        row_has_resource = True
                        resource_col = col_idx

            if row_has_hours and row_has_resource:
                header_row = row_idx
                break
            elif row_has_hours:
                header_row = row_idx
                # Look for resource column
                for col_idx in range(1, min(10, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        cell_str = str(cell_value).lower().strip()
                        if cell_str in ("resource", "name", "team member", "person", "column1"):
                            resource_col = col_idx
                            break
                if resource_col:
                    break

        if not header_row:
            return []

        # Find project names from the row above the header or in the header
        project_row = header_row - 1 if header_row > 1 else header_row

        # Map "Hours" columns to project names
        current_project = None
        for col_idx in range(1, sheet.max_column + 1):
            # Check project name row
            project_cell = sheet.cell(row=project_row, column=col_idx).value
            if project_cell and str(project_cell).strip():
                current_project = str(project_cell).strip()[:50]  # Truncate long names

            # Check if this is an hours column
            header_cell = sheet.cell(row=header_row, column=col_idx).value
            if header_cell:
                header_str = str(header_cell).lower().strip()
                if header_str in ("hours", "hrs", "hour") or header_str.startswith("hours"):
                    if current_project:
                        project_cols[current_project] = col_idx

        if not project_cols:
            return []

        # Parse resource rows and aggregate hours per project
        project_hours = {proj: 0.0 for proj in project_cols}

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            # Skip empty rows
            if resource_col:
                resource_cell = sheet.cell(row=row_idx, column=resource_col).value
                if not resource_cell:
                    continue

            for project, col_idx in project_cols.items():
                hours_cell = sheet.cell(row=row_idx, column=col_idx).value
                if hours_cell is not None:
                    try:
                        hours = float(hours_cell)
                        if hours > 0:
                            project_hours[project] += hours
                    except (ValueError, TypeError) as exc:
                        logger.warning("Non-numeric hours for project %s: %s", project, exc)

        # Create tasks from projects with hours
        tasks = []
        for i, (project, hours) in enumerate(project_hours.items(), 1):
            if hours > 0:
                tasks.append(Task(
                    id=f"PROJ-{i:03d}",
                    name=project,
                    hours=hours,
                    sprint="Sprint 1",  # Default to single sprint
                    tags=["project"],
                ))

        if not tasks:
            return []

        # Create a single sprint with all projects as tasks
        total_hours = sum(t.hours for t in tasks)
        sprint = Sprint(
            name="PI Planning",
            total_hours=total_hours * 1.25,  # Add buffer estimate
            buffer_percentage=self.default_buffer,
            tasks=tasks,
        )

        return [sprint]

    def _identify_sprint_sheets(self, sheet_names: list[str]) -> dict[str, str]:
        """Identify which sheets represent sprints."""
        sprint_sheets = {}
        sprint_pattern = re.compile(r"sprint\s*(\d+)|s(\d+)|iteration\s*(\d+)", re.IGNORECASE)

        for name in sheet_names:
            match = sprint_pattern.search(name)
            if match:
                # Extract sprint number
                sprint_num = next(g for g in match.groups() if g is not None)
                sprint_sheets[name] = f"Sprint {sprint_num}"
            elif name.lower() in ["sprint 1", "sprint 2", "sprint 3", "sprint 4", "sprint 5"]:
                sprint_sheets[name] = name

        return sprint_sheets

    def _parse_sprint_sheet(self, sheet, sprint_name: str) -> Sprint:
        """Parse a single sprint sheet."""
        # Find header row
        headers = {}
        header_row = 1
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            row_values = [str(cell.value).lower() if cell.value else "" for cell in sheet[row_idx]]
            if self._is_header_row(row_values):
                headers = self._map_columns(row_values)
                header_row = row_idx
                break

        if not headers:
            return Sprint(name=sprint_name, total_hours=0)

        # Extract total hours from sheet (look for capacity cells)
        total_hours = self._find_total_hours(sheet)

        # Parse tasks
        tasks = []
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            task = self._parse_task_row(sheet, row_idx, headers, sprint_name)
            if task:
                tasks.append(task)

        sprint = Sprint(
            name=sprint_name,
            total_hours=total_hours if total_hours > 0 else self._calculate_capacity(tasks),
            buffer_percentage=self.default_buffer,
            tasks=tasks,
        )

        return sprint

    def _parse_combined_sheet(self, sheet) -> list[Sprint]:
        """Parse a sheet with multiple sprints indicated by column."""
        # Find header row
        headers = {}
        header_row = 1
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            row_values = [str(cell.value).lower() if cell.value else "" for cell in sheet[row_idx]]
            if self._is_header_row(row_values):
                headers = self._map_columns(row_values)
                header_row = row_idx
                break

        # If no headers found, return empty
        if not headers:
            return []

        # Group tasks by sprint
        sprint_tasks: dict[str, list[Task]] = {}
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            sprint_col = headers.get("sprint")
            if sprint_col is not None:
                sprint_cell = sheet.cell(row=row_idx, column=sprint_col + 1)
                sprint_name = str(sprint_cell.value) if sprint_cell.value else "Unassigned"
            else:
                # No sprint column - put all tasks in "Sprint 1"
                sprint_name = "Sprint 1"

            task = self._parse_task_row(sheet, row_idx, headers, sprint_name)
            if task:
                if sprint_name not in sprint_tasks:
                    sprint_tasks[sprint_name] = []
                sprint_tasks[sprint_name].append(task)

        # Create Sprint objects
        sprints = []
        for sprint_name, tasks in sorted(sprint_tasks.items()):
            sprint = Sprint(
                name=sprint_name,
                total_hours=self._calculate_capacity(tasks) * 1.25,  # Estimate capacity
                buffer_percentage=self.default_buffer,
                tasks=tasks,
            )
            sprints.append(sprint)

        return sprints

    def _is_header_row(self, row_values: list[str]) -> bool:
        """Check if row contains headers."""
        # Count how many known column names we find
        known_columns = (
            self.TASK_ID_COLUMNS
            + self.TASK_NAME_COLUMNS
            + self.HOURS_COLUMNS
            + self.SPRINT_COLUMNS
        )
        matches = sum(1 for val in row_values if any(col in val for col in known_columns))
        return matches >= 2

    def _map_columns(self, row_values: list[str]) -> dict[str, int]:
        """Map column names to indices."""
        headers = {}

        for idx, val in enumerate(row_values):
            val_lower = val.lower().strip()

            if any(col in val_lower for col in self.TASK_ID_COLUMNS):
                headers["task_id"] = idx
            elif any(col in val_lower for col in self.TASK_NAME_COLUMNS):
                headers["task_name"] = idx
            elif any(col in val_lower for col in self.HOURS_COLUMNS):
                headers["hours"] = idx
            elif any(col in val_lower for col in self.SPRINT_COLUMNS):
                headers["sprint"] = idx
            elif any(col in val_lower for col in self.STORY_COLUMNS):
                headers["story"] = idx
            elif any(col in val_lower for col in self.DEPENDENCY_COLUMNS):
                headers["dependencies"] = idx

        return headers

    def _parse_task_row(
        self, sheet, row_idx: int, headers: dict[str, int], sprint_name: str
    ) -> Optional[Task]:
        """Parse a single task row."""
        def get_cell_value(col_name: str) -> str:
            if col_name not in headers:
                return ""
            cell = sheet.cell(row=row_idx, column=headers[col_name] + 1)
            return str(cell.value).strip() if cell.value else ""

        def get_cell_number(col_name: str) -> float:
            if col_name not in headers:
                return 0.0
            cell = sheet.cell(row=row_idx, column=headers[col_name] + 1)
            if cell.value is None:
                return 0.0
            try:
                return float(cell.value)
            except (ValueError, TypeError):
                return 0.0

        task_id = get_cell_value("task_id")
        task_name = get_cell_value("task_name")
        hours = get_cell_number("hours")

        # Skip empty rows
        if not task_name and not task_id:
            return None

        # Generate task ID if not present
        if not task_id:
            task_id = f"TASK-{row_idx:03d}"

        # Parse dependencies
        deps_str = get_cell_value("dependencies")
        dependencies = []
        if deps_str:
            # Split by comma, semicolon, or space
            deps = re.split(r"[,;\s]+", deps_str)
            dependencies = [d.strip() for d in deps if d.strip()]

        story_id = get_cell_value("story") or None

        return Task(
            id=task_id,
            name=task_name or f"Task {task_id}",
            hours=hours,
            sprint=sprint_name,
            story_id=story_id,
            dependencies=dependencies,
        )

    def _find_total_hours(self, sheet) -> float:
        """Find total hours/capacity from sheet."""
        # Look for cells containing "total", "capacity", "available"
        capacity_patterns = ["total hours", "capacity", "available hours", "team capacity"]

        for row in sheet.iter_rows(max_row=min(20, sheet.max_row)):
            for cell in row:
                if cell.value:
                    cell_str = str(cell.value).lower()
                    if any(pattern in cell_str for pattern in capacity_patterns):
                        # Look for number in adjacent cells
                        for offset in [1, 2, -1]:
                            try:
                                adjacent = sheet.cell(row=cell.row, column=cell.column + offset)
                                if adjacent.value and isinstance(adjacent.value, (int, float)):
                                    return float(adjacent.value)
                            except (ValueError, IndexError) as exc:
                                logger.debug("Could not read adjacent cell at offset %d: %s", offset, exc)
                                continue

        return 0.0

    def _calculate_capacity(self, tasks: list[Task]) -> float:
        """Estimate capacity based on task hours."""
        total_task_hours = sum(t.hours for t in tasks)
        # Add 25% buffer to estimate what capacity might have been
        return total_task_hours * 1.25 if total_task_hours > 0 else 80.0
