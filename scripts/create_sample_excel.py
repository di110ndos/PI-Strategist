"""Script to create sample Excel capacity planner."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl")
    exit(1)


def create_sample_capacity_planner():
    """Create a sample capacity planner Excel file."""
    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Sprint 1
    ws1 = wb.active
    ws1.title = "Sprint 1"
    _create_sprint_sheet(ws1, "Sprint 1", 80, [
        ("TASK-301", "Design password reset email template", 4, "STORY-201", ""),
        ("TASK-302", "Implement token generation service", 12, "STORY-201", ""),
        ("TASK-303", "Create password reset API endpoint", 8, "STORY-201", "TASK-302"),
        ("TASK-304", "Build password reset UI", 8, "STORY-201", "TASK-303"),
        ("TASK-305", "Implement MFA TOTP generation", 16, "STORY-202", ""),
        ("TASK-306", "Create MFA setup wizard", 12, "STORY-202", "TASK-305"),
        ("TASK-307", "Build MFA verification flow", 8, "STORY-202", "TASK-305"),
        ("TASK-308", "Create session management dashboard", 16, "STORY-203", ""),
    ], header_font, header_fill, border)

    # Sprint 2
    ws2 = wb.create_sheet("Sprint 2")
    _create_sprint_sheet(ws2, "Sprint 2", 80, [
        ("TASK-309", "Implement session termination API", 8, "STORY-203", ""),
        ("TASK-310", "Add session activity logging", 8, "STORY-203", "TASK-309"),
        ("TASK-311", "Security audit preparation", 16, "STORY-202", "TASK-307"),
        ("TASK-312", "Performance testing", 12, "STORY-201", "TASK-304"),
        ("TASK-313", "Documentation", 8, "STORY-201", ""),
    ], header_font, header_fill, border)

    # Sprint 3
    ws3 = wb.create_sheet("Sprint 3")
    _create_sprint_sheet(ws3, "Sprint 3", 80, [
        ("TASK-314", "Bug fixes and polish", 16, "STORY-201", ""),
        ("TASK-315", "Integration testing", 12, "STORY-202", "TASK-311"),
        ("TASK-316", "User acceptance testing", 8, "STORY-203", "TASK-310"),
        ("TASK-317", "Deployment preparation", 8, "STORY-201", "TASK-312"),
    ], header_font, header_fill, border)

    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    _create_summary_sheet(ws_summary, header_font, header_fill, border)

    # Save
    output_path = Path(__file__).parent.parent / "examples" / "sample-capacity.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Created: {output_path}")


def _create_sprint_sheet(ws, sprint_name, total_hours, tasks, header_font, header_fill, border):
    """Create a sprint sheet with tasks."""
    # Capacity info
    ws["A1"] = "Sprint Capacity"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Total Hours:"
    ws["B2"] = total_hours
    ws["A3"] = "Buffer (20%):"
    ws["B3"] = total_hours * 0.2
    ws["A4"] = "Net Capacity:"
    ws["B4"] = total_hours * 0.8

    # Headers
    headers = ["Task ID", "Task Name", "Hours", "Story", "Dependencies"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Tasks
    for row, task in enumerate(tasks, 7):
        for col, value in enumerate(task, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border

    # Total load
    last_row = 7 + len(tasks)
    ws.cell(row=last_row, column=1, value="Sprint Load:").font = Font(bold=True)
    ws.cell(row=last_row, column=3, value=f"=SUM(C7:C{last_row-1})")

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15


def _create_summary_sheet(ws, header_font, header_fill, border):
    """Create a summary sheet."""
    ws["A1"] = "PI Capacity Summary"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Sprint", "Total Hours", "Buffer", "Net Capacity", "Sprint Load", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    data = [
        ("Sprint 1", 80, 16, 64, 84, "OVER"),
        ("Sprint 2", 80, 16, 64, 52, "OK"),
        ("Sprint 3", 80, 16, 64, 44, "OK"),
    ]

    for row, sprint_data in enumerate(data, 4):
        for col, value in enumerate(sprint_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 6:
                if value == "OVER":
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")

    # Column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15


if __name__ == "__main__":
    create_sample_capacity_planner()
